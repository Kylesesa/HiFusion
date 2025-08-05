import math
import os

import matplotlib.pyplot as plt
import torch
import numpy as np
from PIL import Image
from PIL import ImageEnhance
import torch
import torch.nn.functional as F
from torch.autograd import Variable
from math import exp
from torchvision import transforms
import torch.nn as nn
from torchvision.utils import save_image
from Evaluator import Evaluator
import pandas as pd
import cv2
from openpyxl import load_workbook

loader = transforms.Compose([transforms.ToTensor()])
unloader = transforms.ToPILImage()

def RGB2YCrCb(rgb_image):
    """
    将RGB格式转换为YCrCb格式
    用于中间结果的色彩空间转换中,因为此时rgb_image默认size是[B, C, H, W]
    :param rgb_image: RGB格式的图像数据
    :return: Y, Cr, Cb
    """
    R = rgb_image[:, 0:1]
    G = rgb_image[:, 1:2]
    B = rgb_image[:, 2:3]
    Y = 0.299 * R + 0.587 * G + 0.114 * B
    Cr = (R - Y) * 0.713 + 0.5
    Cb = (B - Y) * 0.564 + 0.5

    Y = Y.clamp(0.0, 1.0)
    Cr = Cr.clamp(0.0, 1.0).detach()
    Cb = Cb.clamp(0.0, 1.0).detach()
    return Y, Cb, Cr

def RGB2Y(img_path):            #rgb_img->YCbCr_img (1, h, w)
    rgb_img = Image.open(img_path)
    if rgb_img.mode == 'RGB':
        ycbcr_img = rgb_img.convert("YCbCr")
        Y = ycbcr_img.split()[0]
        Cb = ycbcr_img.split()[1]
        Cr = ycbcr_img.split()[2]
        tag = True
    else:
        Y = rgb_img
        Cb = rgb_img
        Cr = rgb_img
        tag = False


    return Y, Cb, Cr,tag

# RGB2Y('/home/admin/xkc/RCMCGAN_pytorch/VIFB/VI/fight.jpg')
def YCbCr2RGB(Y, Cb, Cr):
    """
    将YcrCb格式转换为RGB格式
    :param Y:
    :param Cb:
    :param Cr:
    :return:
    """
    ycrcb = torch.cat([Y, Cr, Cb], dim=1)
    B, C, W, H = ycrcb.shape
    im_flat = ycrcb.transpose(1, 3).transpose(1, 2).reshape(-1, 3)
    mat = torch.tensor([[1.0, 1.0, 1.0], [1.403, -0.714, 0.0], [0.0, -0.344, 1.773]]).to(Y.device)
    bias = torch.tensor([0.0 / 255, -0.5, -0.5]).to(Y.device)
    temp = (im_flat + bias).mm(mat)
    out = temp.reshape(B, W, H, C).transpose(1, 3).transpose(2, 3)
    out = out.clamp(0, 1.0)
    return out

# ir_path = '/home/admin/xkc/RCMCGAN_pytorch/VIFB/IR/carLight.jpg'
# vi_path = '/home/admin/xkc/RCMCGAN_pytorch/VIFB/VI/kettle.jpg'
#
# Y, Cb, Cr =RGB2Y(vi_path)


# Y=loader(Y).unsqueeze(0)
# g1 = Y.squeeze().numpy()
# gg1 = Image.fromarray((g1 * 255).astype('uint8'))
# gg1.save('Y.png')

# tensor to PIL Image
def tensor2img(img, is_norm=True):
    img = img.cpu().float().numpy()
    if img.shape[0] == 1:
        img = np.tile(img, (3, 1, 1))
    if is_norm:
        img = (img - np.min(img)) / (np.max(img) - np.min(img))
    img = np.transpose(img, (1, 2, 0)) * 255.0
    return img.astype(np.uint8)

def save_img_single(img, name, is_norm=True):
    img = tensor2img(img, is_norm=True)
    img = Image.fromarray(img)
    img.save(name)


# ssim
def gaussian(window_size, sigma):
    gauss = torch.Tensor([exp(-(x - window_size // 2) ** 2 / float(2 * sigma ** 2)) for x in range(window_size)])
    return gauss / gauss.sum()


def create_window(window_size, channel):
    _1D_window = gaussian(window_size, 1.5).unsqueeze(1)
    _2D_window = _1D_window.mm(_1D_window.t()).float().unsqueeze(0).unsqueeze(0)
    window = Variable(_2D_window.expand(channel, 1, window_size, window_size).contiguous())
    # print(window.shape)  # (1,1,11,11)
    return window


def _ssim(img1, img2, window, window_size, channel, size_average=True):
    mu1 = F.conv2d(img1, window, padding=window_size // 2, groups=channel)
    mu2 = F.conv2d(img2, window, padding=window_size // 2, groups=channel)

    mu1_sq = mu1.pow(2)
    mu2_sq = mu2.pow(2)
    mu1_mu2 = mu1 * mu2

    sigma1_sq = F.conv2d(img1 * img1, window, padding=window_size // 2, groups=channel) - mu1_sq
    sigma2_sq = F.conv2d(img2 * img2, window, padding=window_size // 2, groups=channel) - mu2_sq
    sigma12 = F.conv2d(img1 * img2, window, padding=window_size // 2, groups=channel) - mu1_mu2

    C1 = 0.01 ** 2
    C2 = 0.03 ** 2

    ssim_map = ((2 * mu1_mu2 + C1) * (2 * sigma12 + C2)) / ((mu1_sq + mu2_sq + C1) * (sigma1_sq + sigma2_sq + C2))

    if size_average:
        return ssim_map.mean()
    else:
        return ssim_map
def ssim(img1, img2, window_size=11, size_average=False):#window=11
    (_, channel, _, _) = img1.size()
    window = create_window(window_size, channel)

    if img1.is_cuda:
        window = window.cuda(img1.get_device())
    window = window.type_as(img1)

    return _ssim(img1, img2, window, window_size, channel, size_average)


class SSIM(torch.nn.Module):
    def __init__(self, window_size=11, size_average=True):
        super(SSIM, self).__init__()
        self.window_size = window_size
        self.size_average = size_average
        self.channel = 1
        self.window = create_window(window_size, self.channel)

    def forward(self, img1, img2):
        (_, channel, _, _) = img1.size()

        if channel == self.channel and self.window.data.type() == img1.data.type():
            window = self.window
        else:
            window = create_window(self.window_size, channel)

            if img1.is_cuda:
                window = window.cuda(img1.get_device())
            window = window.type_as(img1)

            self.window = window
            self.channel = channel

        return _ssim(img1, img2, window, self.window_size, channel, self.size_average)


# fused_path = '/home/admin/xkc/RCMCGAN_pytorch/log/log10_8/10/1000/2.png'
# ir_path = '/home/admin/xkc/RCMCGAN_pytorch/test_imgs/IR2.bmp'
# vis_path = '/home/admin/xkc/RCMCGAN_pytorch/test_imgs/VIS2.bmp'
# fused_path2 = '/home/admin/xkc/RCMCGAN_pytorch/log/log5/4/1100/output_image1.png'
# ir_path2 = '/home/admin/xkc/RCMCGAN_pytorch/test_imgs/IR1.bmp'
# vis_path2 = '/home/admin/xkc/RCMCGAN_pytorch/test_imgs/VIS1.bmp'
#
# # fused_img = cv2.imread(fused_path,cv2.IMREAD_GRAYSCALE)
# fused_img = Image.open(fused_path).convert('L')
# ir_img = Image.open(ir_path)
# vis_img = Image.open(vis_path)
# fused_img2 = Image.open(fused_path).convert('L')
# ir_img2 = Image.open(ir_path)
# vis_img2 = Image.open(vis_path2)
# #
# vis1 = loader(vis_img).unsqueeze(0)
# ir1 = loader(ir_img).unsqueeze(0)
# fused1 = loader(fused_img).unsqueeze(0)
# vis2 = loader(vis_img2).unsqueeze(0)
# ir2 = loader(ir_img2).unsqueeze(0)
# fused2 = loader(fused_img2).unsqueeze(0)

# print(ms_ssim(ir1,ir1))
# vis = torch.cat((vis1,vis2),dim=0)
# ir = torch.cat((ir1,ir2),dim=0)
# fused = torch.cat((fused1,fused2),dim=0)
#
# s1 = ssim(vis1, fused2)
# # s2 = ssim(ir, fused)
# has_neg = torch.lt(s1,0).any()    #这一步说明ssim中是存在负数的，这也为什么最终的loss里出现了负数
# if has_neg:
#     print("yes")
# else:
#     print("no")
# print(s1)     #->实验证明，两张一样的图像，ssim输出均为1

# alpha = torch.mean(vis1)
# normalized_data1 = (vis1 - alpha)
# normalized_data1 = F.relu(normalized_data1)
# g1 = normalized_data.squeeze().numpy()
# gg1 = Image.fromarray((g1 * 255).astype('uint8'))
# gg1.save('grad4.png')

# alpha = torch.min(fused1)
# img = fused1 - alpha/3

# alpha = torch.mean(ir1)
# normalized_data = (ir1 - alpha)
# normalized_data = F.relu(normalized_data)
# g1 = cv2.equalizeHist(fused_img)
# bright_map = fused1
# normalized_map = (bright_map - torch.min(bright_map)) / \
#                        (torch.max(bright_map) - torch.min(bright_map))
#
# g1 = img.squeeze().numpy()
# gg1 = Image.fromarray((g1 * 255).astype('uint8'))
# gg1.save('a.png')

# plt.imsave('a.png',g1,cmap='gray')

# B= (normalized_data+normalized_data1)/2
# g1 = B.squeeze().numpy()
# gg1 = Image.fromarray((g1 * 255).astype('uint8'))
# gg1.save('grad6.png')

#---------------------------------------自适应---------------------------------------

class grad_map(nn.Module):
    def __init__(self,window_size):
        super(grad_map, self).__init__()
        self.window_size = window_size
    def grad(self,img):

        b,c,h,w = img.shape
        sobel_x = torch.tensor([[-1, 0, 1],
                                [-2, 0, 2],
                                [-1, 0, 1]], dtype=torch.float32, requires_grad=False).view(1, 1, 3, 3)

        sobel_y = torch.tensor([[-1, -2, -1],
                                [0, 0, 0],
                                [1, 2, 1]], dtype=torch.float32, requires_grad=False).view(1, 1, 3, 3)

        sobel_x = sobel_x.to(img.device)
        sobel_y = sobel_y.to(img.device)

        g_x = F.conv2d(img, sobel_x, padding=1, stride=1,groups=c)
        g_y = F.conv2d(img, sobel_y, padding=1, stride=1,groups=c)
        g = torch.abs(g_x) + torch.abs(g_y)  # (B,C,H,W) 梯度图

        # 拉伸对比度
        normalized_grad = torch.tanh(g)

        reflection_padding = self.window_size // 2
        # reflection_pad = nn.ZeroPad2d(reflection_padding)
        reflection_pad = nn.ReflectionPad2d(reflection_padding)
        padded_g = reflection_pad(normalized_grad)  # 将梯度图边缘填充
        output_tensor = F.avg_pool2d(padded_g, kernel_size=self.window_size, stride=1)  # 得到窗口为9的平均梯度图，但是实际梯度的感受野应为11

        # 对卷积后的梯度再次归一化，这样就不暗了
        output_tensor2 = (output_tensor - torch.min(output_tensor)) / \
                         (torch.max(output_tensor) - torch.min(output_tensor))
        return output_tensor2

    def forward(self, img):
        g_map = self.grad(img)
        return g_map

class bright_map(nn.Module):
    def __init__(self,window_size):
        super(bright_map, self).__init__()
        self.window_size = window_size

    def bright(self, img):
        alpha = torch.mean(img)
        res = (img - alpha)
        # std = torch.std(img) #试过了，好像除与不除，没有什么区别
        # res =res/std
        bright_map = F.relu(res)
        normalized_map = (bright_map - torch.min(bright_map)) / \
                         (torch.max(bright_map) - torch.min(bright_map))

        reflection_padding = self.window_size // 2
        reflection_pad = nn.ReflectionPad2d(reflection_padding)
        # reflection_pad = nn.ZeroPad2d(reflection_padding)
        padded_g = reflection_pad(normalized_map)  # 将梯度图边缘填充
        output_tensor = F.avg_pool2d(padded_g, kernel_size=self.window_size, stride=1)

        output_tensor2 = (output_tensor - torch.min(output_tensor)) / \
                         (torch.max(output_tensor) - torch.min(output_tensor))

        return output_tensor2

    def forward(self, img):
        b_map = self.bright(img)
        return b_map

def grad(img, window_size=3):
    sobel_x = torch.tensor([[-1, 0, 1],
                            [-2, 0, 2],
                            [-1, 0, 1]], dtype=torch.float32,requires_grad=False).view(1, 1, 3, 3)

    sobel_y = torch.tensor([[-1, -2, -1],
                            [0, 0, 0],
                            [1, 2, 1]], dtype=torch.float32,requires_grad=False).view(1, 1, 3, 3)


    sobel_x = sobel_x.to(img.device)
    sobel_y = sobel_y.to(img.device)

    # print(sobel_y)

    g_x = F.conv2d(img, sobel_x, padding=1, stride=1)
    g_y = F.conv2d(img, sobel_y, padding=1, stride=1)
    g = torch.abs(g_x) + torch.abs(g_y) # (B,C,H,W) 梯度图

    #归一化梯度,会影响效果，变暗
    # normalized_grad = (g - torch.min(g)) / \
    #                    (torch.max(g) - torch.min(g))
    normalized_grad = torch.tanh(g)

    reflection_padding = window_size // 2
    reflection_pad = nn.ReflectionPad2d(reflection_padding)
    padded_g = reflection_pad(normalized_grad)  # 将梯度图边缘填充
    output_tensor = F.avg_pool2d(padded_g, kernel_size=window_size, stride=1)   #得到窗口为9的平均梯度图，但是实际梯度的感受野应为11

    # 对卷积后的梯度再次归一化，这样就不暗了
    output_tensor2 = (output_tensor - torch.min(output_tensor)) / \
                      (torch.max(output_tensor) - torch.min(output_tensor))

    return output_tensor2

# def grad_loss(vis, ir, fused_img, window_size = 3):
#     # s3 = torch.exp(ssim(vis, fused_img, window_size))    #SSIM图[-1,1],避免负数的情况（虽然影响不大），通过e映射到[1/e,e]
#     # s4 = torch.exp(ssim(ir, fused_img, window_size))
#
#     s1 = window_l2_norm(vis, fused_img, window_size)
#     s2 = window_l2_norm(ir, fused_img, window_size)
#
#     g_i = torch.exp(grad(ir, window_size))   #首先可以避免分母为0的情况，其次可以拉伸梯度范围[0,1]->[1/e,e]
#     g_v = torch.exp(grad(vis, window_size))
#
#     a = g_v / (g_i + g_v)       #a+b=1
#     b = g_i / (g_i + g_v)
#
#     # ------------改进方式，不让两个特征图抢资源-----------------#
#     '''
#     会报错，在反向传播过程中，导致s1，s2中出现nan，很有可能是a，b的问题
#     '''
#     # a = torch.exp(torch.maximum((g_v - g_i), 0 * (g_i - g_v))) - torch.exp(torch.tensor(-1))
#     # b = torch.exp(torch.maximum((g_i - g_v), 0 * (g_i - g_v))) - torch.exp(torch.tensor(-1))
#
#     # ------------权重的截断+归一化-----------------#
#     a = torch.relu(a - torch.mean(a))
#     # a = torch.tanh(a)/2+0.5
#     a = (a - torch.min(a)) / (torch.max(a) - torch.min(a))
#
#     b = torch.relu(b - torch.mean(b))
#     # b = torch.tanh(b) / 2 + 0.5
#     b = (b - torch.min(b)) / (torch.max(b) - torch.min(b))
#
#     # loss = torch.exp(torch.tensor(1.0)) - torch.mean((a * s1 + b * s2), dim=(2, 3))
#     loss = torch.mean((a * s1 + b * s2), dim=(2, 3))
#     return loss        #[1/e,e]
#     # return a,b

# score = SSIM(vis1,ir1,fused1)
# print(score)
# g1 = grad(vis1,window_size=1)
# g2 = grad(ir1,window_size=1)
# grad1 = (g1+g2)
# max_grad = grad1.max()
# normalized_grad = grad1/max_grad
# g3 = grad(fused1)
# loss = torch.abs((g3-grad1))
# print(loss.shape)
# l2 = torch.norm((g3-g2))
# print(l2)
# print('g:',g)
# has_bigger_than_one = (g1>1).any()    #这一步说明梯度图的范围是[0,1]
# if has_bigger_than_one:
#     print('exist!')
# else:
#     print('no,all smaller than 1')

# gg1 = g1.squeeze().numpy()
# gg1 = Image.fromarray((gg1 * 255).astype('uint8'))
# gg1.save('grad_v2.png')
# gg2 = g2.squeeze().numpy()
# g22 = Image.fromarray((gg2 * 255).astype('uint8'))
# g22.save('grad_i2.png')

#---------------------------------------自适应内容损失---------------------------------------
# def bright(img):
#     '''
#     用11*11的平均核对图像卷积，得到11*11的像素均值
#     :param img:
#     :return:
#     '''
#     window_size = 11
#     mean_kernel = torch.ones(1, 1, window_size, window_size) / (window_size * window_size)
#     mean_kernel = mean_kernel.to(img.device)
#
#     reflection_padding = window_size // 2
#     reflection_pad = nn.ReflectionPad2d(reflection_padding)
#     image = reflection_pad(img)
#
#     mean_map = F.conv2d(image, mean_kernel, padding=0, stride=1)
#
#     return mean_map
#
def bright2(img, window_size):

    alpha = torch.mean(img)
    res = (img - alpha)
    # std = torch.std(img) #试过了，好像除与不除，没有什么区别
    # res =res/std
    bright_map = F.relu(res)
    normalized_map = (bright_map - torch.min(bright_map)) / \
    (torch.max(bright_map) - torch.min(bright_map))

    reflection_padding = window_size // 2
    # reflection_pad = nn.ReflectionPad2d(reflection_padding)
    reflection_pad = nn.ZeroPad2d(reflection_padding)
    padded_g = reflection_pad(normalized_map)  # 将梯度图边缘填充
    output_tensor = F.avg_pool2d(padded_g, kernel_size=window_size, stride=1)

    output_tensor2 = (output_tensor - torch.min(output_tensor)) / \
                     (torch.max(output_tensor) - torch.min(output_tensor))

    return output_tensor2


def window_l1_norm(img1, img2):
    '''
    用于求两张图象的l1范数，|x1|+|x2|+|x3|+...+|x121|
    :param img1:
    :param img2:
    :return:
    '''
    window_size = 11
    stride = 1

    reflection_padding = window_size // 2
    reflection_pad = nn.ReflectionPad2d(reflection_padding)
    image1 = reflection_pad(img1).to(img1.device)
    image2 = reflection_pad(img2).to(img2.device)

    res_map = image1 - image2

    unfolded = res_map.unfold(2, window_size, stride).unfold(3, window_size, stride)

    output_tensor = (torch.abs(unfolded)).sum(dim=(4, 5))
    return output_tensor


# def window_l2_norm(img1, img2, window_size=3):
#     stride = 1
#     eps = 1e-8
#     reflection_padding = window_size // 2
#     reflection_pad = nn.ReflectionPad2d(reflection_padding)
#     image1 = reflection_pad(img1).to(img1.device)
#     image2 = reflection_pad(img2).to(img2.device)
#
#     res_map = image1 - image2
#
#     unfolded = res_map.unfold(2, window_size, stride).unfold(3, window_size, stride)
#     # output_tensor = torch.norm(unfolded+eps,dim=[4,5])
#     output_tensor = torch.sqrt(torch.sum(torch.square(unfolded), dim=[4, 5])+eps)
#     # print(output_tensor.shape)
#     return output_tensor

class window_l2_norm(nn.Module):
    def __init__(self,window_size):
        super(window_l2_norm, self).__init__()
        self.window_size = window_size

    def _window_l2_norm(self, img1, img2):
        stride = 1
        eps = 1e-8
        reflection_padding = self.window_size // 2
        reflection_pad = nn.ReflectionPad2d(reflection_padding)
        image1 = reflection_pad(img1).to(img1.device)
        image2 = reflection_pad(img2).to(img2.device)

        res_map = image1 - image2

        unfolded = res_map.unfold(2, self.window_size, stride).unfold(3, self.window_size, stride)
        # output_tensor = torch.norm(unfolded+eps,dim=[4,5])
        output_tensor = torch.sqrt(torch.sum(torch.square(unfolded), dim=[4, 5])+eps)
        # print(output_tensor.shape)
        return output_tensor

    def forward(self, img1,img2):
        norm_map = self._window_l2_norm(img1,img2)
        return norm_map

class window_ssim(nn.Module):
    def __init__(self,window_size):
        super(window_ssim, self).__init__()
        self.window_size = window_size
        self.ssim = SSIM(window_size=self.window_size,size_average=False)

    def forward(self, img1, img2):
        output_tensor = self.ssim(img1, img2)
        # print(output_tensor.shape)
        return output_tensor

# g1 = grad(vis1,window_size=3)       #3,7,15
# g2 = grad(ir1,window_size=3)
# grad1 = (g1+g2)/2
# max_grad = torch.max(grad1)
# min_grad = grad1.min()
# normalized_grad = (grad1-min_grad)/(max_grad-min_grad)
# grad1 = F.relu(grad1)
# g11 = g1.squeeze().numpy()
# gg1 = Image.fromarray((g11 * 255).astype('uint8'))
# gg1.save('b4_v.png')

# a = window_l2_norm(vis1,ir1)
# g22 = g2.squeeze().numpy()
# gg2= Image.fromarray((g22 * 255).astype('uint8'))
# gg2.save('b4_i.png')
# has_neg = torch.lt(a, 1).any()  # 这一步说明ssim中是存在负数的，这也为什么最终的loss里出现了负数
# if has_neg:
#     print("yes")
# else:
#     print("no")


# print(a.shape)
# print('a',a)
# g = a.squeeze().numpy()
# gg = Image.fromarray((g * 255).astype('uint8'))
# gg.save('e.png')

# def bright_loss(vis, ir, fused_img, window_size=61):
#     eps = 1e-8
#     c1 = window_l2_norm(vis, fused_img, window_size)
#     c2 = window_l2_norm(ir, fused_img, window_size)
#     # s1 = torch.exp(ssim(vis, fused_img))  # SSIM图[-1,1],避免负数的情况（虽然影响不大），通过e映射到[1/e,e]
#     # s2 = torch.exp(ssim(ir, fused_img))
#
#
#     b_i = torch.exp(bright2(ir, window_size))  # 拉开二者的差距，同时避免分母为0
#     b_v = torch.exp(bright2(vis, window_size))
#     # b_i = bright2(ir)
#     # b_v = bright2(vis)
#     # b_map = b_v + b_i
#     # B = (b_map - torch.min(b_map)) / (torch.max(b_map) - torch.min(b_map))
#
#     B = b_v + b_i
#     a = b_v / (B)
#     b = b_i / (B)
#
#     #------------改进方式，不让两个特征图抢资源-----------------#
#     # a = torch.exp(F.relu(b_v - b_i))-torch.exp(torch.tensor(-1))
#     # b = torch.exp(F.relu(b_i - b_v))-torch.exp(torch.tensor(-1))
#
#     # ------------权重的截断+归一化-----------------#
#     a=torch.relu(a-torch.mean(a))
#     a = (a - torch.min(a)) / (torch.max(a) - torch.min(a))
#
#     b = torch.relu(b - torch.mean(b))
#     b = (b - torch.min(b)) / (torch.max(b) - torch.min(b))
#
#
#     loss = torch.mean((a * c1 + b * c2), dim=(2, 3))
#     # score = torch.mean((a * s1 + b * s2), dim=(2, 3))
#     return loss
#     # return a,b


# def weight_map(vis, ir, window_size=3):
#     eps = 1e-8
#     b_i = torch.exp(bright2(ir, window_size))
#     b_v = torch.exp(bright2(vis, window_size))
#     # b_i = bright2(ir, window_size)
#     # b_v = bright2(vis, window_size)
#
#     B = b_v + b_i
#     a1 = b_v / (B)
#     b1 = b_i / (B)
#
#     g_i = torch.exp(grad(ir, window_size))
#     g_v = torch.exp(grad(vis, window_size))
#     # g_i = grad(ir, window_size)
#     # g_v = grad(vis, window_size)
#
#     G = g_v + g_i
#     a2 = g_v / (G)
#     b2 = g_i / (G)
#
#     a = (a1 + a2)/2
#     b = (b1 + b2)/2
#
#     # i=b_i+g_i
#     # v=b_v+g_v
#     #
#     # a = v/(i+v+eps)
#     # b = i/(i+v+eps)
#
#
#     # ------------权重的截断+归一化-----------------#
#     map_a = torch.relu(a-torch.mean(a))
#     map_v = (map_a - torch.min(map_a)) / (torch.max(map_a) - torch.min(map_a))
#     # a2 = torch.tanh(a2)
#     map_b = torch.relu(b - torch.mean(b))
#     map_i = (map_b - torch.min(map_b)) / (torch.max(map_b) - torch.min(map_b))
#     # b2 = torch.tanh(b2)
#
#     return map_v, map_i     #a,b是全局特征。a2,b2是细节特征


# def con_loss(vis, ir, fused_img, window_size=3):
#
#     c1 = window_l2_norm(vis, fused_img, window_size)
#     c2 = window_l2_norm(ir, fused_img, window_size)
#
#     # c3 = 1-ssim(vis, fused_img,size_average=False,window_size=window_size)
#     # c4 = 1-ssim(ir, fused_img,size_average=False,window_size=window_size)
#     # c3 = 1-ms_ssim(vis, fused_img, size_average=False,win_size=window_size)
#     # c4 = 1-ms_ssim(ir, fused_img, size_average=False,win_size=window_size)
#
#     a2, b2 = weight_map(vis, ir, window_size)
#     # loss = torch.mean((a * (c1+100*c3) + b * (c2+100*c4)), dim=(2, 3))
#     # loss1 = torch.mean((a  *  c3 + b  *  c4), dim=(2, 3))
#     loss2 = torch.mean((a2 *  c1 + b2 *  c2), dim=(2, 3))
#     # return 100*loss1+loss2
#     return loss2

class con_loss(nn.Module):
    def __init__(self,window_size):
        super(con_loss, self).__init__()
        self.window_size = window_size
        self.bright = bright_map(self.window_size)
        self.grad = grad_map(self.window_size)
        self.window_l2_norm =window_l2_norm(self.window_size)
        # self.window_ssim = window_ssim(self.window_size)

    def weight_map(self, vis, ir):#先各自求权重，再相加得到权重图

        b_i = torch.exp(self.bright(ir))
        b_v = torch.exp(self.bright(vis))

        B = b_v + b_i
        a1 = b_v / (B)
        b1 = b_i / (B)

        g_i = torch.exp(self.grad(ir))
        g_v = torch.exp(self.grad(vis))

        G = g_v + g_i
        a2 = g_v / (G)
        b2 = g_i / (G)

        a = (0.4*a1 + 0.6*a2)
        b = (0.4*b1 + 0.6*b2)

        # ------------权重的截断+归一化-----------------#
        # map_a = a
        map_a = torch.relu(a - torch.mean(a))
        map_v = (map_a - torch.min(map_a)) / (torch.max(map_a) - torch.min(map_a))

        # map_b = b
        map_b = torch.relu(b - torch.mean(b))
        map_i = (map_b - torch.min(map_b)) / (torch.max(map_b) - torch.min(map_b))

        return map_v, map_i

    def weight_map2(self, vis, ir):#先算信息图，再求权重

        b_i = self.bright(ir)
        b_v = self.bright(vis)

        g_i = self.grad(ir)
        g_v = self.grad(vis)

        i = torch.exp(b_i + g_i)
        v = torch.exp(b_v + g_v)

        a = v/(i+v)
        b = i/(i+v)

        # ------------权重的截断+归一化-----------------#
        # map_a = a
        map_a = torch.relu(a - torch.mean(a))
        map_v = (map_a - torch.min(map_a)) / (torch.max(map_a) - torch.min(map_a))

        # map_b = b
        map_b = torch.relu(b - torch.mean(b))
        map_i = (map_b - torch.min(map_b)) / (torch.max(map_b) - torch.min(map_b))

        return map_v, map_i

    def _con_loss(self,vis, ir, fused_img):
        # ------------计算窗口F范数-----------------#
        n1 = self.window_l2_norm(vis, fused_img)
        n2 = self.window_l2_norm(ir, fused_img)

        # ------------计算窗口ssim-----------------#
        # s1 = 1 - self.window_ssim(vis, fused_img)
        # s2 = 1 - self.window_ssim(ir, fused_img)

        c1 = n1
        c2 = n2

        a, b = self.weight_map(vis, ir)
        #-----------------
        loss = torch.mean((a * c1 + b * c2), dim=(2, 3))
        #--------------
        # tensor1=fused_img-vis
        # tensor2=fused_img-ir
        # eps = 1e-3
        #
        # loss1 = torch.sqrt(torch.sum(torch.square(tensor1), dim=[2, 3])+eps)
        # loss2 = torch.sqrt(torch.sum(torch.square(tensor2), dim=[2, 3])+eps)
        # loss= loss2
        # print(loss)
        #----------------

        return loss

    def forward(self, vis, ir, fused_img):
        loss_map = self._con_loss(vis, ir, fused_img)
        return loss_map
# v=plt.imread(vis_path)
# i=plt.imread(ir_path)
# c3 = 1-ssim(vis1, ir1, size_average=False,window_size=3)
# print(c3.shape)

# c=con_loss(window_size=61)
# a,b= c.weight_map2(vis1, ir1)
# # print(a.shape)
# g1 = a.squeeze().numpy()
# gg1 = Image.fromarray((g1 * 255).astype('uint8'))
# plt.imshow(gg1,cmap='jet')
# plt.axis('off')
# plt.savefig('w_v.png',pad_inches=0)
#
# b1 = b.squeeze().numpy()
# bb1 = Image.fromarray((b1 * 255).astype('uint8'))
# plt.imshow(bb1,cmap='jet')
# plt.axis('off')
# plt.savefig('w_i.png',pad_inches=0)


# 用于将文件夹中的RGB图片转换为灰度图
# image_path = '/home/admin/xkc/RCMCGAN_pytorch/TNO/vi'
# i = 0
# for filename in os.listdir(image_path):
#     if not filename.startswith('.'):
#         img_path = os.path.join(image_path, filename)
#         save_path = os.path.join(image_path, filename)
#         img = Image.open(img_path)
#
#         if img.mode != 'L':
#             img = img.convert('L')
#             img.save(save_path)
#             i += 1
#             print(i)


#---------------------------------------CBAM---------------------------------------
# 通道注意力
class ChannelAttention(nn.Module):
    def __init__(self, in_planes, ratio=16):
        super(ChannelAttention, self).__init__()
        # 平均池化
        self.avg_pool = nn.AdaptiveAvgPool2d(1)
        # 最大池化
        self.max_pool = nn.AdaptiveMaxPool2d(1)

        # MLP  除以16是降维系数
        self.fc1 = nn.Conv2d(in_planes, in_planes // ratio, 1, bias=False)  # kernel_size=1
        self.relu1 = nn.ReLU()
        self.fc2 = nn.Conv2d(in_planes // ratio, in_planes, 1, bias=False)

        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        avg_out = self.fc2(self.relu1(self.fc1(self.avg_pool(x))))
        max_out = self.fc2(self.relu1(self.fc1(self.max_pool(x))))
        # 结果相加
        out = avg_out + max_out
        return self.sigmoid(out)


# 空间注意力
class SpatialAttention(nn.Module):
    def __init__(self, kernel_size=7):
        super(SpatialAttention, self).__init__()
        # 声明卷积核为 3 或 7
        assert kernel_size in (3, 7), 'kernel size must be 3 or 7'
        # 进行相应的same padding填充
        padding = 3 if kernel_size == 7 else 1

        self.conv1 = nn.Conv2d(2, 1, kernel_size, padding=padding, bias=False)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        avg_out = torch.mean(x, dim=1, keepdim=True)  # 平均池化
        max_out, _ = torch.max(x, dim=1, keepdim=True)  # 最大池化
        # 拼接操作
        x = torch.cat([avg_out, max_out], dim=1)
        x = self.conv1(x)  # 7x7卷积填充为3，输入通道为2，输出通道为1
        return self.sigmoid(x)

class CBAM(nn.Module):
    def __init__(self, in_planes, ratio=1, kernel_size=7):
        super(CBAM, self).__init__()
        self.channel_att = ChannelAttention(in_planes,ratio)
        self.spatial_att = SpatialAttention(kernel_size)

    def forward(self,x):
        out = self.channel_att(x)*x
        out = self.spatial_att(out)*out
        return out

# model = CBAM(3)
# input = torch.ones((1,3,224,224))
# output = model(input)
# print(input.shape)
# print(output.shape)


#---------------------------------------ECA---------------------------------------

class ECA(nn.Module):
    def __init__(self, channel, b=1, gamma=2):
        super(ECA, self).__init__()
        kernel_size = int(abs(math.log(channel,2)+b)/gamma)
        kernel_size = kernel_size if kernel_size%2 else kernel_size+1

        self.avg_pool = nn.AdaptiveAvgPool2d(1)
        self.conv = nn.Conv1d(1,1,kernel_size=kernel_size, padding=(kernel_size-1)//2, bias=False)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        y=self.avg_pool(x)
        y = self.conv(y.squeeze(-1).transpose(-1,-2)).transpose(-1,-2).unsqueeze(-1)
        y = self.sigmoid(y)
        out = x*y.expand_as(x)

        return out

# model = ECA(3)
# input = torch.ones((1,3,224,224))
# output = model(input)
# print(input.shape)
# print(output.shape)

#---------------------------------------wgan-gp---------------------------------------
def gradient_penalty(D, real, fake, device):
    b, c, h, w = real.shape
    epsilon = torch.rand((b, 1, 1, 1), requires_grad=True).repeat(1, c, h, w).to(device)
    new_img = real * epsilon + fake * (1 - epsilon)

    mix_score = D(new_img)

    gradient = torch.autograd.grad(
        inputs=new_img,
        outputs=mix_score,
        grad_outputs=torch.ones_like(mix_score),
        create_graph=True,
        retain_graph=True,
    )[0]

    gradient = gradient.view(gradient.shape[0],-1)
    gradient_norm = gradient.norm(2,dim=1)
    penalty = torch.mean((gradient_norm-1)**2)

    return penalty

#---------------------------------------LeakyRelu---------------------------------------
def lrelu(x, leak=0.2):
    return torch.maximum(x, leak*x)
#---------------------------------------Grad_Loss---------------------------------------



#---------------------------------------TV_Loss---------------------------------------
def tv(img):
    tv_x = torch.pow((img[:, :, :, :-1] - img[:, :, :, 1:]), 2)
    tv_y = torch.pow((img[:, :, :-1, :] - img[:, :, 1:, :]), 2)
    tv_loss = tv_x + tv_y
    return tv_loss

#---------------------------------------get_metrics---------------------------------------#
def image_read_cv2(path, mode='RGB'):
    img_BGR = cv2.imread(path).astype('float32')
    assert mode == 'RGB' or mode == 'GRAY' or mode == 'YCrCb', 'mode error'
    if mode == 'RGB':
        img = cv2.cvtColor(img_BGR, cv2.COLOR_BGR2RGB)
    elif mode == 'GRAY':
        img = np.round(cv2.cvtColor(img_BGR, cv2.COLOR_BGR2GRAY))
    elif mode == 'YCrCb':
        img = cv2.cvtColor(img_BGR, cv2.COLOR_BGR2YCrCb)
    return img

def get_metrics(ori_img_folder,img_path):
    metric_result = np.zeros((10))
    for img_name in os.listdir(os.path.join(ori_img_folder, "ir")):
        ir = image_read_cv2(os.path.join(ori_img_folder, "ir", img_name), 'GRAY')
        vi = image_read_cv2(os.path.join(ori_img_folder, "vi", img_name), 'GRAY')
        fi = image_read_cv2(os.path.join(img_path, img_name), 'GRAY')
        metric_result += np.array([Evaluator.EN(fi), Evaluator.SD(fi),
                                   Evaluator.SF(fi),Evaluator.VIFF(fi, ir, vi),
                                   Evaluator.AG(fi),Evaluator.SCD(fi, ir, vi),
                                   Evaluator.MI(fi, ir, vi), Evaluator.Qabf(fi, ir, vi),
                                   Evaluator.SSIM(fi, ir, vi),Evaluator.CC(fi, ir, vi)])

    metric_result /= len(os.listdir(img_path))
    return metric_result

def save_metrics_to_excel(metrics, file_path):
    df = pd.DataFrame([metrics], columns=[f'Metric_{i+1}' for i in range(len(metrics))])
    if not os.path.isfile(file_path):
        df.to_excel(file_path, index=False, header=True)
    else:
        book = load_workbook(file_path)
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            writer.book = book
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
            for sheetname in writer.sheets:
                df.to_excel(writer, sheet_name=sheetname, index=False, header=False, startrow=writer.sheets[sheetname].max_row)


